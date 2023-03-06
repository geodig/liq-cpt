# -*- coding: utf-8 -*-
"""
Created on Fri Mar  3 14:04:54 2023

@author: YOGB
"""
import streamlit as st
# from streamlit_extras.app_logo import add_logo
import folium
from streamlit_folium import st_folium
import openpyxl
import pandas
import utm
from module.cptliqsett import CPT_LIQ

st.set_page_config(page_title="WBLIQ-CPT"
                    ,layout="wide"
                   )

# st.sidebar.image("https://climateapp.nl/img/about/logo-witteveenbos.png", use_column_width='auto')


@st.cache_data
def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv().encode('utf-8')


tab1, tab2, tab3 = st.tabs(["Graph", "Table", "Info"])

with st.sidebar.expander("**INPUT FILES**"):
    uploaded_files = st.file_uploader("Upload CPT file(s):", type=["xlsx"], accept_multiple_files=True)
    if uploaded_files:
        listfile = [i.name for i in uploaded_files]

        with st.sidebar.form("**INPUT PARAMETERS**"):
            
            select = st.selectbox("Select CPT profile:", options=listfile)
            index = listfile.index(select)
            gwd = st.slider("Ground water depth (m):", min_value = 0.0, max_value = 10.0, step = 0.1)
            PGA = st.slider("PGA (g):", min_value = 0.0, max_value = 1.0, step = 0.01)
            Mw = st.slider("Magnitude (Mw):", min_value = 3.5, max_value = 9.5, step = 0.1)
            method = st.selectbox("Liquefaction severity method:", options=["LDCR (Hutabarat and Bray, 2021)",
                                                                            "LPI (Iwasaki et al, 1978)",
                                                                            "LSN (Tonkin and Taylor, 2013)"])
            submit = st.form_submit_button("Submit")
            
        wb = openpyxl.load_workbook(uploaded_files[index], data_only=True)
        sh = wb["general"]
        sh_header, sh_value = [],[]
        for i in range(sh.max_row):
            sh_header.append(sh.cell(i+1,1).value)
            sh_value.append([sh.cell(i+1,2).value])
        
        sh_comb = {}
        for i in range(len(sh_header)):
            sh_comb[sh_header[i]] = sh_value[i]
        sh_df = pandas.DataFrame(sh_comb)
        sh_df = pandas.DataFrame.transpose(sh_df)
        
        X,Y,UTM_z,UTM_h = sh_value[2][0],sh_value[3][0],sh_value[7][0],sh_value[8][0]
        label = sh_value[1][0]
        if UTM_h == "N":
            hemis = True
        elif UTM_h == "S":
            hemis = False
        latlon = utm.to_latlon(X, Y, UTM_z, northern=hemis)
        lat = latlon[0]
        lon = latlon[1]
        
        location = [lat, lon]
        m = folium.Map(location=location, zoom_start=16, tiles="CartoDB positron")
        folium.Marker([lat, lon], popup=label, icon=folium.Icon(color='salmon')).add_to(m)
        
        with tab3:
            col1, col2 = st.columns(2)
            with col1:
                mapfol = st_folium(m, height=600, width=700)
            with col2:
                st.table(sh_df)
        
        if submit:
            if method == "LDCR (Hutabarat and Bray, 2021)":
                cpt = CPT_LIQ(file=wb, Mw=Mw, PGA=PGA, gwd=gwd)
                result = cpt.LDCR()
                fig = result[0]
                df = result[1]
                LD = result[2]
                CR = result[3]
                category = result[4]
                
            elif method == "LPI (Iwasaki et al, 1978)":
                cpt = CPT_LIQ(file=wb, Mw=Mw, PGA=PGA, gwd=gwd)
                result = cpt.LPI()
                fig = result[0]
                df = result[1]
                LPI = result[2]
                category = result[3]
            
            elif method == "LSN (Tonkin and Taylor, 2013)":
                cpt = CPT_LIQ(file=wb, Mw=Mw, PGA=PGA, gwd=gwd)
                result = cpt.LSN()
                fig = result[0]
                df = result[1]
                LSN = result[2]
                category = result[3]
            
            st.session_state["result_fig"] = fig
            st.session_state["result_df"] = df
        
        fig = st.session_state["result_fig"]
        df = st.session_state["result_df"]

        with tab1:
            st.pyplot(fig)
        with tab2:
            st.write(df)
            csv = convert_df(df)
            st.download_button(
                label="Download CSV",
                data=csv,
                file_name='result_%s.csv'%(select),
                mime='text/csv')

    else:
        tab1.warning("Please upload files from the sidebar.",icon="⚠️")
        tab2.warning("Please upload files from the sidebar.",icon="⚠️")
        tab3.warning("Please upload files from the sidebar.",icon="⚠️")


with st.sidebar.expander("**REFERENCES**"):
    st.markdown("""
                Hutabarat, D. and J. D. Bray, 2021, "_Estimating the severity of liquefaction
ejecta using the cone penetration test_", Journal of Geotechnical and 
Geoenvironmental Engineering, 148(3). 

Iwasaki, T., F. Tatsuoka, K. Tokida, and S. Yasuda, 1978, “_A practical
method for assessing soil liquefaction potential based on case studies
at various sites in Japan_”, in Proc., 2nd Int. Earthquake Microzonation
Conf., 885–896, Washington, DC: National Science Foundation.

Tonkin and Taylor, 2013, _Liquefaction vulnerability study_, Auckland,
New Zealand: Tonkin and Taylor.

Zhang, G., P. Robertson, and R. W. Brachman, 2002, “_Estimating
liquefaction-induced ground settlements from CPT for level ground_”,
Can. Geotech. J. 39 (5): 1168–1180.
                """)







